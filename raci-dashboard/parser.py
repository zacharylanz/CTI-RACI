"""
Flexible RACI spreadsheet parser.

Auto-detects layout from .xlsx or .csv files without requiring
a specific column order or naming convention.

Handles:
  - Standard RACI (R/A/C/I single letters)
  - Extended variants: RASCI, RACI-VS, DACI, RAPID
  - Full-word values ("Responsible", "Accountable", etc.)
  - Multi-value cells ("R/A", "R,A") — keeps highest responsibility
  - Inline category headers (rows with name but no RACI)
  - Explicit category columns
  - Merged title/banner rows
  - Sub-header rows with full role names
  - Maturity scales 0-5 (auto-normalized from 0-10 or percentages)
  - Summary/aggregate sections at bottom of sheet
  - Multiple sheets (auto-selects best RACI sheet)
  - CSV with various encodings (UTF-8, Latin-1, CP1252)
  - Transposed layouts (roles as rows)
"""

import csv
import io
import os
import re
RACI_VALUES = {'R', 'A', 'C', 'I'}

# Extended RACI variants mapped to standard R/A/C/I
RACI_EXTENDED = {
    # Standard
    'R': 'R', 'A': 'A', 'C': 'C', 'I': 'I',
    # RASCI: S(upportive) → C
    'S': 'C',
    # RACI-VS: V(erify) → C, S(ign-off) → A
    'V': 'C',
    # DACI: D(river) → R, A(pprover) → A, C(ontributor) → C, I(nformed) → I
    'D': 'R',
    # RAPID: R(ecommend) → C, A(gree) → A, P(erform) → R, I(nput) → C, D(ecide) → A
    'P': 'R',
    # Common extras
    'X': 'R',  # "X" mark often means responsible
    'O': 'R',  # "Owner"
    'L': 'R',  # "Lead"
}

# Full-word RACI values → letter (case-insensitive matching)
RACI_FULLWORDS = {
    'responsible': 'R', 'accountable': 'A', 'consulted': 'C', 'informed': 'I',
    'supportive': 'C', 'support': 'C',
    'driver': 'R', 'approver': 'A', 'contributor': 'C',
    'perform': 'R', 'recommend': 'C', 'input': 'C', 'decide': 'A',
    'lead': 'R', 'owner': 'R', 'participant': 'C',
    'verify': 'C', 'sign-off': 'A', 'sign off': 'A',
    'yes': 'R', 'y': 'R',  # Some sheets use Y/N
}

# Responsibility priority (higher = more responsible)
RACI_PRIORITY = {'R': 4, 'A': 3, 'C': 2, 'I': 1}

MATURITY_RANGE = {0, 1, 2, 3, 4, 5}

ROLE_PALETTE = [
    "#4ae0b0", "#e0a040", "#6090e0", "#a0b8d0",
    "#e06080", "#80d0d0", "#d080e0", "#c0c060",
    "#50b890", "#d09060", "#7080d0", "#b0c8e0",
    "#d070a0", "#60c0b0", "#c090d0", "#b0b070",
]
CATEGORY_PALETTE = [
    "#8090CC", "#50C890", "#90C850", "#B888CC",
    "#C8A050", "#A080C0", "#C89850", "#6898B8", "#58A8C0",
    "#7888B8", "#60B880", "#A0B850", "#C898C0",
    "#B8A060", "#9078B0", "#D0A858", "#5890A8",
]

# Column classification keywords
HEADER_KEYWORDS = {
    'name': [
        'capability', 'name', 'activity', 'task', 'function', 'process',
        'item', 'deliverable', 'work package', 'work item', 'responsibility',
        'action', 'objective', 'requirement', 'service', 'control',
    ],
    'description': [
        'desc', 'description', 'details', 'notes', 'comment', 'explanation',
        'definition', 'summary', 'scope',
    ],
    'category': [
        'category', 'domain', 'area', 'group', 'pillar', 'section',
        'phase', 'stream', 'workstream', 'department', 'team', 'module',
        'tower', 'theme', 'bucket', 'cluster',
    ],
    'maturity_now': [
        'now', 'current', 'maturity', 'level', 'baseline', 'as-is', 'as is',
        'actual', 'present', 'today',
    ],
    'maturity_target': [
        'target', 'tgt', 'future', 'goal', 'projected', 'to-be', 'to be',
        'desired', 'planned', 'expected', 'objective',
    ],
    'delta': [
        'delta', 'uplift', 'gap', 'Δ', 'diff', 'difference', 'variance',
        'change', 'improvement',
    ],
    'status': [
        'status', 'state', 'fill', 'progress', 'completion',
    ],
    'priority': [
        'priority', 'prio', 'importance', 'urgency', 'rank', 'weight',
    ],
    'id': [
        'id', '#', 'no', 'number', 'ref', 'reference', 'code', 'key',
    ],
}

# Keywords that indicate maturity_target vs maturity_now
TARGET_KEYWORDS = [
    'target', 'tgt', 'future', 'goal', 'projected', 'to-be', 'to be',
    'desired', 'planned', 'expected', 'with',
]
UNFILLED_KEYWORDS = ['open', 'unfilled', 'vacant', '★', 'tbd', 'tbc', 'hire', 'needed', 'new']

# Patterns that indicate a row is a summary/aggregate (not a real capability)
SUMMARY_KEYWORDS = [
    'average', 'avg', 'total', 'sum', 'count', 'mean', 'median',
    'grand total', 'subtotal', 'sub-total', 'summary',
    'category average', 'section total',
]

# Patterns that indicate a category is a footer/summary section (not real data)
SUMMARY_CATEGORY_KEYWORDS = [
    'average', 'avg', 'total', 'sum', 'count', 'legend', 'key',
    'summary', 'appendix', 'reference', 'notes', 'glossary',
    'responsible (r)', 'accountable (a)', 'consulted (c)', 'informed (i)',
    'raci legend', 'raci key', 'raci count', 'count by role',
]


def _cell_str(val):
    """Convert cell value to stripped string."""
    if val is None:
        return ''
    return str(val).strip()


def _cell_raw(val):
    """Convert cell value to stripped string preserving case."""
    if val is None:
        return ''
    return str(val).strip()


def _normalize_raci(val):
    """
    Normalize a cell value to a standard RACI letter.

    Handles:
      - Single letters: R, A, C, I, S, D, V, X, etc.
      - Full words: "Responsible", "Accountable", etc.
      - Multi-value: "R/A", "R,A", "R & A" → picks highest priority
      - Extended variants: RASCI, DACI, RAPID → maps to R/A/C/I
      - Common marks: X, Y → R

    Returns the RACI letter or '' if not recognized.
    """
    s = _cell_str(val)
    if not s:
        return ''

    upper = s.upper()

    # Single standard RACI letter
    if upper in RACI_VALUES:
        return upper

    # Single extended letter
    if upper in RACI_EXTENDED:
        return RACI_EXTENDED[upper]

    # Full word match (case-insensitive)
    lower = s.lower()
    if lower in RACI_FULLWORDS:
        return RACI_FULLWORDS[lower]

    # Multi-value: split on / , & and pick highest priority
    parts = re.split(r'[/,&\s]+', upper)
    mapped = []
    for p in parts:
        p = p.strip()
        if p in RACI_EXTENDED:
            mapped.append(RACI_EXTENDED[p])
    if mapped:
        return max(mapped, key=lambda x: RACI_PRIORITY.get(x, 0))

    # Partial word match (starts with a known word)
    for word, letter in RACI_FULLWORDS.items():
        if lower.startswith(word):
            return letter

    return ''


def _is_raci(val):
    """Check if value can be interpreted as a RACI assignment."""
    return _normalize_raci(val) != ''


def _is_maturity_number(val, scale_max=5):
    """Check if value is a maturity number (0 to scale_max)."""
    s = _cell_str(val)
    if s == '':
        return False
    # Strip % sign
    s = s.rstrip('%').strip()
    try:
        n = float(s)
        return 0 <= n <= max(scale_max, 5)
    except (ValueError, TypeError):
        return False


def _detect_maturity_scale(values):
    """
    Detect the maturity scale from a list of numeric values.
    Returns (scale_max, is_percentage).
    Common scales: 0-5, 0-10, 0-100 (percentage).
    """
    nums = []
    for v in values:
        s = _cell_str(v).rstrip('%').strip()
        try:
            nums.append(float(s))
        except (ValueError, TypeError):
            continue
    if not nums:
        return 5, False
    max_val = max(nums)
    if max_val > 10:
        return 100, True  # percentage scale
    elif max_val > 5:
        return 10, False
    return 5, False


def _normalize_maturity(val, scale_max=5):
    """Convert a maturity value to 0-5 scale."""
    s = _cell_str(val).rstrip('%').strip()
    try:
        n = float(s)
    except (ValueError, TypeError):
        return None
    if scale_max == 100:
        return round(n / 20)  # 0-100 → 0-5
    elif scale_max == 10:
        return round(n / 2)   # 0-10 → 0-5
    return round(n)


def _make_short_code(label):
    """Derive a short code from a role label."""
    label = label.strip()
    # If already short enough (<=5 chars), use uppercase
    if len(label) <= 5:
        return label.upper()
    # Try initials from multi-word labels
    words = re.findall(r'[A-Z][a-z]*|[a-z]+', label)
    if len(words) >= 2:
        initials = ''.join(w[0] for w in words if w[0].isalpha())
        if 2 <= len(initials) <= 5:
            return initials.upper()
    # Try uppercase consonants
    consonants = re.sub(r'[aeiou\s\W]', '', label, flags=re.IGNORECASE)
    if len(consonants) >= 3:
        return consonants[:4].upper()
    # Fallback: first 4 characters
    return label[:4].upper()


def _make_id(label):
    """Create a snake_case id from label."""
    s = re.sub(r'[^a-zA-Z0-9\s]', '', label)
    s = re.sub(r'\s+', '_', s.strip())
    return s.lower()


def _detect_unfilled(header_text):
    """Detect if a role header indicates an unfilled position."""
    ht = header_text.lower()
    return any(kw in ht for kw in UNFILLED_KEYWORDS)


def _is_summary_row(name_val):
    """Check if a row name indicates it's a summary/aggregate row."""
    lower = name_val.lower().strip()
    return any(kw in lower for kw in SUMMARY_KEYWORDS)


def _is_summary_category(cat_name):
    """Check if a category name indicates it's a footer/summary section."""
    lower = cat_name.lower().strip()
    return any(kw in lower for kw in SUMMARY_CATEGORY_KEYWORDS)


def _strip_category_numbering(name):
    """Remove leading numbers/bullets from category names: '1. Strategy' → 'Strategy'."""
    # Strip "1.", "1)", "1 -", "a.", "a)", bullet chars
    s = re.sub(r'^[\d]+[.):\-]\s*', '', name.strip())
    s = re.sub(r'^[a-zA-Z][.)]\s*', '', s)
    s = re.sub(r'^[•●○◦▪▸►→–—]\s*', '', s)
    return s.strip() or name.strip()


def _load_xlsx(filepath, sheet_name=None):
    """Load an Excel file and return rows as list of lists + sheet name used."""
    from openpyxl import load_workbook
    wb = load_workbook(filepath, data_only=True)
    if sheet_name:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            )
    else:
        ws = _pick_best_sheet(wb)
    used_sheet = ws.title

    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))

    # Handle merged cells: if the workbook has merged cell ranges,
    # openpyxl with data_only fills the top-left, rest are None.
    # Re-open without data_only to get merged ranges, then fill values.
    try:
        wb2 = load_workbook(filepath)
        ws2 = wb2[used_sheet]
        for merge_range in ws2.merged_cells.ranges:
            min_row = merge_range.min_row
            max_row = merge_range.max_row
            min_col = merge_range.min_col
            max_col = merge_range.max_col
            # Get the value from the top-left cell
            val = rows[min_row - 1][min_col - 1] if min_row - 1 < len(rows) else None
            for r in range(min_row - 1, min(max_row, len(rows))):
                for c in range(min_col - 1, min(max_col, len(rows[r]) if r < len(rows) else 0)):
                    rows[r][c] = val
    except Exception:
        pass  # If merge detection fails, proceed with what we have

    return rows, used_sheet


def _pick_best_sheet(wb):
    """
    Auto-select the best sheet in a multi-sheet workbook.
    Scores each sheet by how much it looks like a RACI matrix.
    Falls back to active sheet if no clear winner.
    """
    if len(wb.sheetnames) == 1:
        return wb.active

    best_score = -1
    best_ws = wb.active

    for name in wb.sheetnames:
        ws = wb[name]
        score = 0
        name_lower = name.lower()

        # Sheet name hints
        if 'raci' in name_lower:
            score += 50
        if 'maturity' in name_lower:
            score += 20
        if any(kw in name_lower for kw in ['responsibility', 'assignment', 'matrix']):
            score += 30
        # Penalize obvious non-RACI sheets
        if any(kw in name_lower for kw in ['chart', 'graph', 'pivot', 'lookup', 'config', 'template', 'instruction', 'readme', 'cover']):
            score -= 50

        # Sample data for RACI content
        raci_count = 0
        cell_count = 0
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri > 30:
                break
            for val in row:
                s = _cell_str(val).upper()
                if s:
                    cell_count += 1
                if s in RACI_VALUES:
                    raci_count += 1

        if cell_count > 0:
            raci_density = raci_count / cell_count
            score += int(raci_density * 100)

        if score > best_score:
            best_score = score
            best_ws = ws

    return best_ws


def _load_csv(filepath):
    """Load a CSV file with auto-encoding detection."""
    # Try encodings in order of likelihood
    for encoding in ['utf-8-sig', 'utf-8', 'cp1252', 'latin-1', 'iso-8859-1']:
        try:
            rows = []
            with open(filepath, 'r', encoding=encoding) as f:
                # Sniff delimiter
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
                except csv.Error:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                for row in reader:
                    rows.append(row)
            if rows:
                return rows, 'CSV'
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError("Could not read CSV file with any supported encoding")


def _find_header_row(rows, max_scan=25):
    """
    Find the header row: first row with 4+ DISTINCT non-empty cells.
    Skips merged title rows where all cells have the same value.
    Also skips rows that look like metadata (date, author, version, etc.)
    """
    for i, row in enumerate(rows[:max_scan]):
        non_empty = [_cell_str(c) for c in row if _cell_str(c) != '']
        if len(non_empty) >= 4 and len(set(non_empty)) >= 3:
            # Extra check: skip rows where values are mostly numeric (data, not headers)
            numeric_count = sum(1 for v in non_empty if re.match(r'^[\d.,%]+$', v))
            if numeric_count / len(non_empty) < 0.6:
                return i
    # Fallback: first row with 3+ distinct non-empty cells
    for i, row in enumerate(rows[:max_scan]):
        non_empty = [_cell_str(c) for c in row if _cell_str(c) != '']
        if len(non_empty) >= 3 and len(set(non_empty)) >= 2:
            return i
    # Last resort: first row with any content
    for i, row in enumerate(rows[:max_scan]):
        if any(_cell_str(c) != '' for c in row):
            return i
    return 0


def _skip_subheader_rows(rows, header_idx):
    """
    Skip sub-header rows immediately after the header.
    A sub-header row has 3+ filled cells but contains no RACI values and no maturity numbers.
    Rows with only 1-2 filled cells are NOT sub-headers (likely category headers).
    Returns (skip_count, list_of_subheader_rows).
    """
    skip_count = 0
    subheader_rows = []
    for i in range(header_idx + 1, min(header_idx + 5, len(rows))):
        row = rows[i]
        non_empty = [c for c in row if _cell_str(c) != '']
        if len(non_empty) < 3:
            break  # Not a sub-header row
        has_raci = any(_is_raci(c) for c in row)
        has_maturity = any(_is_maturity_number(c) for c in row)
        if has_raci or has_maturity:
            break  # Data row, stop skipping
        subheader_rows.append(row)
        skip_count += 1
    return skip_count, subheader_rows


def _classify_columns(headers, data_rows):
    """
    Classify each column by inspecting header text and data values.
    Returns a dict: col_index -> classification string.
    """
    num_cols = len(headers)
    classifications = {}
    # Collect data stats per column
    col_stats = {}
    for ci in range(num_cols):
        values = []
        for row in data_rows:
            if ci < len(row):
                v = _cell_str(row[ci])
                if v:
                    values.append(v)
        col_stats[ci] = values

    # First pass: classify by header keywords
    header_lower = [_cell_str(h).lower() for h in headers]

    for ci, hl in enumerate(header_lower):
        if not hl:
            continue
        # Check delta/skip first (to exclude before maturity checks)
        for kw in HEADER_KEYWORDS['delta']:
            if kw.lower() in hl:
                classifications[ci] = 'delta'
                break
        if ci in classifications:
            continue
        # Check status
        for kw in HEADER_KEYWORDS['status']:
            if kw in hl:
                classifications[ci] = 'status'
                break
        if ci in classifications:
            continue
        # Check priority (skip these columns, not useful for RACI)
        for kw in HEADER_KEYWORDS['priority']:
            if kw in hl:
                classifications[ci] = 'priority'
                break
        if ci in classifications:
            continue
        # Check ID columns (skip these) — use exact/prefix match to avoid
        # false positives like 'no' matching 'now'
        if hl in ('#', 'id', 'no', 'no.', 'ref', 'ref.', 'key'):
            classifications[ci] = 'id'
            continue
        for kw in HEADER_KEYWORDS['id']:
            # Require exact match, or keyword followed by non-alpha (e.g., "ref #")
            if hl == kw or re.match(rf'^{re.escape(kw)}[\s._#\-]+', hl) or re.match(rf'^[\s._#\-]+{re.escape(kw)}$', hl):
                classifications[ci] = 'id'
                break

    # Second pass: classify by data patterns for unclassified columns
    name_col_found = False
    desc_col_found = False
    for ci in range(num_cols):
        if ci in classifications:
            continue
        hl = header_lower[ci]
        values = col_stats.get(ci, [])
        total = len(values)
        if total == 0:
            classifications[ci] = 'empty'
            continue

        # Check RACI pattern: count values that normalize to RACI
        raci_count = sum(1 for v in values if _normalize_raci(v) != '')
        raci_pct = raci_count / total if total > 0 else 0

        # Check maturity pattern: >40% are numbers in valid range
        mat_count = sum(1 for v in values if _is_maturity_number(v, 100))
        mat_pct = mat_count / total if total > 0 else 0

        # Check if it's a text column with repeating values (category)
        unique_ratio = len(set(v.lower() for v in values)) / total if total > 0 else 1

        # Check description (long text)
        avg_len = sum(len(v) for v in values) / total if total > 0 else 0

        # Check if column is purely numeric (IDs, sequence numbers)
        numeric_count = sum(1 for v in values if re.match(r'^[\d]+\.?[\d]*$', v))
        numeric_pct = numeric_count / total if total > 0 else 0

        # RACI detection: either high RACI percentage or header suggests roles
        if raci_pct > 0.3:
            # But avoid misclassifying numeric columns with values 1-5
            # that happen to overlap with maturity. RACI columns should have
            # mostly single letters.
            letter_count = sum(1 for v in values if len(v.strip()) <= 3)
            if letter_count / total > 0.3:
                classifications[ci] = 'raci'
                continue

        if mat_pct > 0.4 and numeric_pct > 0.4:
            # Detect the scale for this column
            scale_max, _ = _detect_maturity_scale(values)
            # Distinguish now vs target by header keywords
            is_target = any(kw in hl for kw in TARGET_KEYWORDS)
            existing_mat = [c for c, t in classifications.items() if t == 'maturity_now']
            if is_target or existing_mat:
                classifications[ci] = 'maturity_target'
            else:
                classifications[ci] = 'maturity_now'
        elif any(kw in hl for kw in HEADER_KEYWORDS['description']):
            classifications[ci] = 'description'
            desc_col_found = True
        elif any(kw in hl for kw in HEADER_KEYWORDS['category']):
            classifications[ci] = 'category'
        elif any(kw in hl for kw in HEADER_KEYWORDS['name']):
            classifications[ci] = 'name'
            name_col_found = True
        elif not name_col_found and avg_len > 3 and unique_ratio > 0.5 and numeric_pct < 0.5:
            # Likely the name column: first text column with diverse values
            if unique_ratio < 0.3 and total > 5:
                classifications[ci] = 'category'
            else:
                classifications[ci] = 'name'
                name_col_found = True
        elif not desc_col_found and avg_len > 30 and unique_ratio > 0.7:
            # Long, diverse text → probably description
            classifications[ci] = 'description'
            desc_col_found = True
        elif unique_ratio < 0.3 and total > 3:
            classifications[ci] = 'category'
        elif numeric_pct > 0.8:
            # Purely numeric column that isn't maturity — skip it
            classifications[ci] = 'numeric_skip'
        else:
            classifications[ci] = 'unknown'

    # If no name column found, assign the first unclassified text column
    if not any(v == 'name' for v in classifications.values()):
        for ci in range(num_cols):
            if classifications.get(ci) in ('unknown', None):
                classifications[ci] = 'name'
                break
        # If still none, use column 0
        if not any(v == 'name' for v in classifications.values()):
            classifications[0] = 'name'

    return classifications


def _detect_transposed(rows, header_idx, classifications):
    """
    Detect if the layout is transposed (roles as rows, capabilities as columns).
    In a transposed layout, the first column after the header contains role names,
    and the header row contains capability names.

    Returns True if layout appears transposed.
    """
    raci_cols = [ci for ci, t in classifications.items() if t == 'raci']
    name_cols = [ci for ci, t in classifications.items() if t == 'name']

    # If we found RACI columns normally, it's not transposed
    if len(raci_cols) >= 2:
        return False

    # Check if the first column values look like role names
    # and most row data is RACI letters
    if not rows or header_idx >= len(rows):
        return False

    headers = rows[header_idx]
    data_rows = rows[header_idx + 1:]

    if len(data_rows) < 2:
        return False

    # Count RACI values across all data cells (excluding first column)
    total_cells = 0
    raci_cells = 0
    for row in data_rows[:20]:
        for ci in range(1, min(len(row), len(headers))):
            v = _cell_str(row[ci])
            if v:
                total_cells += 1
                if _normalize_raci(v):
                    raci_cells += 1

    if total_cells > 0 and raci_cells / total_cells > 0.3:
        # Also check: few rows (roles) but many columns (capabilities)
        if len(data_rows) < 20 and len(headers) > len(data_rows) * 2:
            return True

    return False


def _parse_transposed(rows, header_idx):
    """
    Parse a transposed RACI layout where roles are rows and capabilities are columns.
    Returns the same data structure as the normal parse path.
    """
    headers = rows[header_idx]
    data_rows = rows[header_idx + 1:]

    # First column is role names, rest are capabilities
    roles = []
    role_items = {}  # role_index -> {cap_name: raci_val}

    for i, row in enumerate(data_rows):
        if not row:
            continue
        role_name = _cell_raw(row[0])
        if not role_name:
            continue
        # Skip summary rows
        if _is_summary_row(role_name):
            continue

        role_id = _make_id(role_name)
        short = _make_short_code(role_name)
        is_unfilled = _detect_unfilled(role_name)
        roles.append({
            'id': role_id,
            'label': role_name,
            'short': short,
            'color': ROLE_PALETTE[i % len(ROLE_PALETTE)],
            'status': 'unfilled' if is_unfilled else 'filled',
        })

        for ci in range(1, len(row)):
            cap_name = _cell_raw(headers[ci]) if ci < len(headers) else ''
            if not cap_name:
                continue
            val = _normalize_raci(row[ci])
            if val:
                if cap_name not in role_items:
                    role_items[cap_name] = {}
                role_items[cap_name][role_id] = val

    # Build categories (single "General" category for transposed)
    items = []
    for cap_name, assignments in role_items.items():
        item = {'name': cap_name}
        item.update(assignments)
        items.append(item)

    categories = [{
        'name': 'General',
        'color': CATEGORY_PALETTE[0],
        'items': items,
    }] if items else []

    total_capabilities = len(items)
    orphaned = []
    for item in items:
        has_r = any(item.get(r['id']) == 'R' for r in roles)
        if not has_r:
            orphaned.append(f"General > {item['name']}")

    meta = {
        'filename': '',
        'sheet': '',
        'role_count': len(roles),
        'category_count': len(categories),
        'capability_count': total_capabilities,
        'orphaned_capabilities': orphaned,
        'zero_r_roles': [],
        'has_maturity': False,
        'column_classifications': {},
        'layout': 'transposed',
    }

    return {'roles': roles, 'categories': categories, 'meta': meta}


def parse_file(filepath, sheet_name=None):
    """
    Parse a RACI spreadsheet and return structured data.

    Returns:
        dict with 'roles', 'categories', 'meta'
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.csv':
        rows, sheet_used = _load_csv(filepath)
    elif ext in ('.xlsx', '.xls'):
        rows, sheet_used = _load_xlsx(filepath, sheet_name)
    else:
        raise ValueError(f"Unsupported file format: {ext}. Use .xlsx or .csv")

    if not rows:
        raise ValueError("File is empty or unreadable")

    # Step 1: Find header row
    header_idx = _find_header_row(rows)
    headers = rows[header_idx]

    # Normalize header length
    max_cols = max(len(r) for r in rows)
    headers = headers + [None] * (max_cols - len(headers))
    for i, row in enumerate(rows):
        rows[i] = row + [None] * (max_cols - len(row))

    # Step 2: Skip sub-header rows
    skip, subheader_rows = _skip_subheader_rows(rows, header_idx)
    data_start = header_idx + 1 + skip
    data_rows = rows[data_start:]

    # Step 3: Classify columns
    classifications = _classify_columns(headers, data_rows)

    # Step 3.5: Check for transposed layout
    if _detect_transposed(rows, header_idx, classifications):
        result = _parse_transposed(rows, header_idx)
        result['meta']['filename'] = os.path.basename(filepath)
        result['meta']['sheet'] = sheet_used
        return result

    # Step 4: Extract role info from RACI columns
    raci_cols = {ci: _cell_raw(headers[ci]) for ci, t in classifications.items() if t == 'raci'}
    if not raci_cols:
        raise ValueError(
            "No RACI columns detected. Ensure your spreadsheet has columns "
            "where values are R, A, C, or I (or extended variants like RASCI).\n"
            "Supported layouts:\n"
            "  Capability | Role1 | Role2 | ... (with R/A/C/I values)\n"
            "  Task | PM | Dev | QA | Design\n"
            "Also supports: full words (Responsible, Accountable, ...),\n"
            "  multi-value cells (R/A), and RASCI/DACI/RAPID variants."
        )

    # Build sub-header lookup: column index -> full name from sub-header row
    subheader_labels = {}
    for sub_row in subheader_rows:
        for ci in raci_cols:
            if ci < len(sub_row):
                val = _cell_raw(sub_row[ci])
                if val and len(val) > 1:
                    subheader_labels[ci] = val

    roles = []
    for i, (ci, label) in enumerate(sorted(raci_cols.items())):
        # Prefer sub-header full name for display, but keep short header as 'short'
        full_label = subheader_labels.get(ci, label)
        short_code = _make_short_code(label)
        # Use short header text if it's actually short (abbreviation)
        if len(label) <= 6 and label == label.upper():
            short_code = label
        role_id = _make_id(full_label)
        # Check unfilled status in both header and sub-header
        is_unfilled = _detect_unfilled(label) or _detect_unfilled(full_label)
        roles.append({
            'id': role_id,
            'label': full_label,
            'short': short_code,
            'color': ROLE_PALETTE[i % len(ROLE_PALETTE)],
            'status': 'unfilled' if is_unfilled else 'filled',
            'col_index': ci,
        })

    # Step 5: Find name, category, description, maturity columns
    name_col = next((ci for ci, t in classifications.items() if t == 'name'), None)
    cat_col = next((ci for ci, t in classifications.items() if t == 'category'), None)
    desc_col = next((ci for ci, t in classifications.items() if t == 'description'), None)
    mat_now_col = next((ci for ci, t in classifications.items() if t == 'maturity_now'), None)
    mat_tgt_col = next((ci for ci, t in classifications.items() if t == 'maturity_target'), None)

    # Detect maturity scale from data
    mat_scale = 5
    if mat_now_col is not None:
        mat_values = [_cell_str(row[mat_now_col]) for row in data_rows
                      if mat_now_col < len(row) and _cell_str(row[mat_now_col])]
        if mat_tgt_col is not None:
            mat_values += [_cell_str(row[mat_tgt_col]) for row in data_rows
                           if mat_tgt_col < len(row) and _cell_str(row[mat_tgt_col])]
        mat_scale, _ = _detect_maturity_scale(mat_values)

    # Step 6: Build categories and items
    categories_dict = {}  # name -> {color, items}
    current_category = 'General'

    for row in data_rows:
        # Skip completely empty rows
        non_empty = [c for c in row if _cell_str(c) != '']
        if not non_empty:
            continue

        name_val = _cell_raw(row[name_col]) if name_col is not None and name_col < len(row) else ''

        # Check if this is a category header row (inline category detection):
        # Name column has a value but ALL RACI columns are empty.
        # Must run BEFORE summary skip so that "CATEGORY AVERAGES" etc.
        # become their own category (filtered later by has_any_raci).
        raci_values_in_row = [_cell_str(row[ci]) for ci in raci_cols.keys() if ci < len(row)]
        all_raci_empty = all(v == '' for v in raci_values_in_row)

        if name_val and all_raci_empty and not cat_col:
            # This is an inline category header
            current_category = _strip_category_numbering(name_val)
            continue

        # Skip summary/aggregate rows (after category detection)
        if name_val and _is_summary_row(name_val):
            continue

        # Skip rows with no name
        if not name_val:
            continue

        # Determine category
        if cat_col is not None and cat_col < len(row):
            cat_val = _cell_raw(row[cat_col])
            if cat_val:
                current_category = _strip_category_numbering(cat_val)

        # Build item
        item = {'name': name_val}

        if desc_col is not None and desc_col < len(row):
            desc_val = _cell_raw(row[desc_col])
            if desc_val:
                item['desc'] = desc_val

        # RACI assignments — use _normalize_raci for extended format support
        for role in roles:
            ci = role['col_index']
            if ci < len(row):
                val = _normalize_raci(row[ci])
                if val:
                    item[role['id']] = val

        # Maturity (normalized to 0-5)
        if mat_now_col is not None and mat_now_col < len(row):
            normalized = _normalize_maturity(row[mat_now_col], mat_scale)
            if normalized is not None:
                item['now'] = normalized
        if mat_tgt_col is not None and mat_tgt_col < len(row):
            normalized = _normalize_maturity(row[mat_tgt_col], mat_scale)
            if normalized is not None:
                item['tgt'] = normalized

        # Add to category
        if current_category not in categories_dict:
            categories_dict[current_category] = {'items': []}
        categories_dict[current_category]['items'].append(item)

    # Assign category colors (skip summary/footer sections)
    role_ids = [r['id'] for r in roles]
    categories = []
    color_idx = 0
    for cat_name, cat_data in categories_dict.items():
        if not cat_data['items']:
            continue
        # Skip categories whose names indicate summary/footer content
        if _is_summary_category(cat_name):
            continue
        # Check if at least one item in this category has a RACI value
        has_any_raci = any(
            any(item.get(rid) in RACI_VALUES for rid in role_ids)
            for item in cat_data['items']
        )
        if not has_any_raci:
            continue  # Skip summary/aggregate sections
        categories.append({
            'name': cat_name,
            'color': CATEGORY_PALETTE[color_idx % len(CATEGORY_PALETTE)],
            'items': cat_data['items'],
        })
        color_idx += 1

    # Remove col_index from roles (internal use only)
    for role in roles:
        del role['col_index']

    # Build validation report
    total_capabilities = sum(len(c['items']) for c in categories)
    orphaned = []
    for cat in categories:
        for item in cat['items']:
            has_r = any(item.get(r['id']) == 'R' for r in roles)
            if not has_r:
                orphaned.append(f"{cat['name']} > {item['name']}")

    zero_r_roles = []
    for role in roles:
        r_count = sum(
            1 for cat in categories for item in cat['items']
            if item.get(role['id']) == 'R'
        )
        if r_count == 0:
            zero_r_roles.append(role['label'])

    col_report = {
        ci: {'header': _cell_raw(headers[ci]) if ci < len(headers) else '',
             'classification': t}
        for ci, t in sorted(classifications.items())
        if t not in ('empty', 'delta', 'priority', 'id', 'numeric_skip', 'unknown')
    }

    meta = {
        'filename': os.path.basename(filepath),
        'sheet': sheet_used,
        'role_count': len(roles),
        'category_count': len(categories),
        'capability_count': total_capabilities,
        'orphaned_capabilities': orphaned,
        'zero_r_roles': zero_r_roles,
        'has_maturity': mat_now_col is not None,
        'maturity_scale': mat_scale,
        'column_classifications': col_report,
        'layout': 'standard',
    }

    return {
        'roles': roles,
        'categories': categories,
        'meta': meta,
    }


def parse_file_from_bytes(file_bytes, filename, sheet_name=None):
    """Parse RACI data from in-memory bytes (for upload endpoint)."""
    import tempfile
    ext = os.path.splitext(filename)[1].lower()
    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        return parse_file(tmp_path, sheet_name)
    finally:
        os.unlink(tmp_path)
